import io
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, LongTable
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment
from decimal import Decimal
from datetime import datetime, date
from io import BytesIO


# ====================================================
# 🩷 GENERADOR DE PDF — FULL WIDTH — COLOR ROSADO
# ====================================================
def generar_reporte_pdf(nombre_archivo, titulo, columnas, filas):

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=10,
        rightMargin=10,
        topMargin=20,
        bottomMargin=20
    )

    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"<b>{titulo}</b>", styles["Title"]))
    story.append(Spacer(1, 18))

    tabla_data = [columnas] + filas
    num_cols = len(columnas)
    total_width = 11 * inch
    col_width = total_width / num_cols
    col_widths = [col_width] * num_cols

    tabla = LongTable(tabla_data, colWidths=col_widths, repeatRows=1)

    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#EC4899")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#FCE7F3")),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor("#4A044E")),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor("#DB2777")),
    ]))

    story.append(tabla)
    doc.build(story)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'
    return response


# ====================================================
# ⚪ EXCEL — FORMATO REAL .xlsx
# ====================================================
def generar_reporte_excel(data_para_reporte, interpretacion):

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    prompt_usado = interpretacion.get('prompt', 'Reporte')
    sheet.title = prompt_usado[:30]

    if not data_para_reporte:
        sheet.cell(row=1, column=1).value = "No se encontraron datos."
    else:
        headers = list(data_para_reporte[0].keys())
        formatted_headers = [h.split('__')[-1].replace('_', ' ').capitalize() for h in headers]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        for col_num, header_title in enumerate(formatted_headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 22

        for row_num, row_data in enumerate(data_para_reporte, 2):
            for col_num, header_key in enumerate(headers, 1):
                value = row_data.get(header_key, None)
                cell = sheet.cell(row=row_num, column=col_num)

                if isinstance(value, Decimal):
                    cell.value = float(value)
                elif isinstance(value, datetime):
                    cell.value = value.replace(tzinfo=None)
                elif isinstance(value, date):
                    cell.value = value
                else:
                    cell.value = str(value) if value else ''

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"reporte_{prompt_usado.lower().replace(' ', '_')[:30]}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
